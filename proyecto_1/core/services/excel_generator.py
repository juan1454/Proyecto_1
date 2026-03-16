from openpyxl import load_workbook
from pathlib import Path
from django.conf import settings
from datetime import datetime

#generador del nombre del archivo

def generar_nombre_excel(numero_id):
    """
    Genera un nombre único para el documento
    """

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    return f"excel_{numero_id}_{timestamp}.xlsx"


def generar_excel_desde_plantilla(plantilla, datos, numero_id, usuario):

    ruta_plantilla = Path(settings.MEDIA_ROOT) / plantilla.archivo.name
    wb = load_workbook(ruta_plantilla)

    # recorrer TODAS las hojas
    for ws in wb.worksheets:

        for row in ws.iter_rows():
            for cell in row:

                if cell.value is not None:
                    texto = str(cell.value)

                    for key, value in datos.items():
                        if key in texto:
                            texto = texto.replace(key, str(value))

                    cell.value = texto

    # carpeta salida
    salida_dir = Path(settings.MEDIA_ROOT) / "documentos_generados" / usuario
    salida_dir.mkdir(parents=True, exist_ok=True)

    nombre_excel = generar_nombre_excel(numero_id)
    ruta_salida = salida_dir / nombre_excel

    wb.save(ruta_salida)

    return ruta_salida


